"""
Export router - CSV and Excel payroll-ready exports.
"""
from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from typing import Optional, List
import csv
import io

from core.database import get_database
from core.security import get_current_user, TokenData
from core.exceptions import ForbiddenException, BadRequestException
from middleware.permissions import require_permission
from models.role import has_permission, get_role_data_scope, DataScope
from models.base import generate_uuid, utc_now
from services.audit_service import audit_log

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter(prefix="/exports", tags=["Exports"])


async def get_export_data(
    db,
    branch_id: Optional[str],
    start_date: str,
    end_date: str,
    scope: DataScope,
    user_branch_id: Optional[str],
    user_team_id: Optional[str],
    status_filter: List[str] = None
):
    """Get data for export based on filters and scope."""
    
    # Build match filter
    match_filter = {
        "date": {"$gte": start_date, "$lte": end_date}
    }
    
    # Apply scope restrictions
    if scope == DataScope.BRANCH:
        match_filter["branch_id"] = user_branch_id
    elif scope == DataScope.TEAM:
        match_filter["team_id"] = user_team_id
    elif branch_id and scope == DataScope.ALL:
        match_filter["branch_id"] = branch_id
    
    if status_filter:
        match_filter["status"] = {"$in": status_filter}
    
    # Aggregation pipeline
    pipeline = [
        {"$match": match_filter},
        {
            "$lookup": {
                "from": "users",
                "localField": "user_id",
                "foreignField": "id",
                "as": "user"
            }
        },
        {"$unwind": {"path": "$user", "preserveNullAndEmptyArrays": True}},
        {
            "$lookup": {
                "from": "branches",
                "localField": "branch_id",
                "foreignField": "id",
                "as": "branch"
            }
        },
        {"$unwind": {"path": "$branch", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 0,
                "employee_id": "$user.employee_id",
                "first_name": "$user.first_name",
                "last_name": "$user.last_name",
                "email": "$user.email",
                "branch_code": "$branch.code",
                "branch_name": "$branch.name",
                "date": 1,
                "clock_in_time": "$clock_in.timestamp",
                "clock_out_time": "$clock_out.timestamp",
                "total_hours": 1,
                "regular_hours": 1,
                "overtime_hours": 1,
                "break_minutes": 1,
                "status": 1,
                "is_manual_entry": 1
            }
        },
        {"$sort": {"date": 1, "employee_id": 1}}
    ]
    
    return await db.time_entries.aggregate(pipeline).to_list(10000)


def format_time(timestamp: Optional[str]) -> str:
    """Format timestamp for export."""
    if not timestamp:
        return ""
    try:
        dt = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
        return dt.strftime("%H:%M")
    except:
        return timestamp


@router.get("/timesheet/csv")
async def export_timesheet_csv(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = Query(None, description="Filter by status: completed,approved"),
    current_user: TokenData = Depends(get_current_user)
):
    """Export timesheet data as CSV."""
    db = get_database()
    
    if not has_permission(current_user.role, "reports.export"):
        raise ForbiddenException("Cannot export data")
    
    # Default date range (current pay period - 2 weeks)
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start = datetime.now(timezone.utc) - timedelta(days=14)
        start_date = start.strftime("%Y-%m-%d")
    
    scope = get_role_data_scope(current_user.role)
    status_filter = status.split(",") if status else ["completed", "approved"]
    
    # Get data
    data = await get_export_data(
        db, branch_id, start_date, end_date, scope,
        current_user.branch_id, current_user.team_id, status_filter
    )
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row
    writer.writerow([
        "Employee ID",
        "First Name",
        "Last Name",
        "Email",
        "Branch Code",
        "Branch Name",
        "Date",
        "Clock In",
        "Clock Out",
        "Total Hours",
        "Regular Hours",
        "Overtime Hours",
        "Break (min)",
        "Status",
        "Manual Entry"
    ])
    
    # Data rows
    for row in data:
        writer.writerow([
            row.get("employee_id", ""),
            row.get("first_name", ""),
            row.get("last_name", ""),
            row.get("email", ""),
            row.get("branch_code", ""),
            row.get("branch_name", ""),
            row.get("date", ""),
            format_time(row.get("clock_in_time")),
            format_time(row.get("clock_out_time")),
            row.get("total_hours", 0),
            row.get("regular_hours", 0),
            row.get("overtime_hours", 0),
            row.get("break_minutes", 0),
            row.get("status", ""),
            "Yes" if row.get("is_manual_entry") else "No"
        ])
    
    # Audit log
    await audit_log(
        db=db,
        actor_id=current_user.user_id,
        actor_email=current_user.email,
        actor_role=current_user.role,
        action="report.export",
        target_type="timesheet",
        target_id="csv",
        metadata={
            "start_date": start_date,
            "end_date": end_date,
            "branch_id": branch_id,
            "record_count": len(data)
        }
    )
    
    # Return CSV file
    output.seek(0)
    filename = f"timesheet_{start_date}_to_{end_date}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/payroll/csv")
async def export_payroll_csv(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: TokenData = Depends(get_current_user)
):
    """Export payroll-ready summary CSV (aggregated by employee)."""
    db = get_database()
    
    if not has_permission(current_user.role, "reports.export"):
        raise ForbiddenException("Cannot export data")
    
    # Default date range
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start = datetime.now(timezone.utc) - timedelta(days=14)
        start_date = start.strftime("%Y-%m-%d")
    
    scope = get_role_data_scope(current_user.role)
    
    # Build match filter
    match_filter = {
        "date": {"$gte": start_date, "$lte": end_date},
        "status": {"$in": ["completed", "approved"]}
    }
    
    if scope == DataScope.BRANCH:
        match_filter["branch_id"] = current_user.branch_id
    elif branch_id and scope == DataScope.ALL:
        match_filter["branch_id"] = branch_id
    
    # Aggregate by employee
    pipeline = [
        {"$match": match_filter},
        {
            "$group": {
                "_id": "$user_id",
                "total_hours": {"$sum": {"$ifNull": ["$total_hours", 0]}},
                "regular_hours": {"$sum": {"$ifNull": ["$regular_hours", 0]}},
                "overtime_hours": {"$sum": {"$ifNull": ["$overtime_hours", 0]}},
                "days_worked": {"$addToSet": "$date"},
                "entries_count": {"$sum": 1}
            }
        },
        {
            "$lookup": {
                "from": "users",
                "localField": "_id",
                "foreignField": "id",
                "as": "user"
            }
        },
        {"$unwind": "$user"},
        {
            "$lookup": {
                "from": "branches",
                "localField": "user.branch_id",
                "foreignField": "id",
                "as": "branch"
            }
        },
        {"$unwind": {"path": "$branch", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 0,
                "employee_id": "$user.employee_id",
                "first_name": "$user.first_name",
                "last_name": "$user.last_name",
                "email": "$user.email",
                "branch_code": "$branch.code",
                "branch_name": "$branch.name",
                "rate_tier": "$user.hourly_rate_tier",
                "total_hours": {"$round": ["$total_hours", 2]},
                "regular_hours": {"$round": ["$regular_hours", 2]},
                "overtime_hours": {"$round": ["$overtime_hours", 2]},
                "days_worked": {"$size": "$days_worked"},
                "entries_count": 1
            }
        },
        {"$sort": {"branch_code": 1, "employee_id": 1}}
    ]
    
    data = await db.time_entries.aggregate(pipeline).to_list(1000)
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row - Payroll ready format
    writer.writerow([
        "Employee ID",
        "First Name",
        "Last Name",
        "Email",
        "Branch Code",
        "Branch Name",
        "Rate Tier",
        "Days Worked",
        "Regular Hours",
        "Overtime Hours",
        "Total Hours",
        "Pay Period Start",
        "Pay Period End"
    ])
    
    # Data rows
    for row in data:
        writer.writerow([
            row.get("employee_id", ""),
            row.get("first_name", ""),
            row.get("last_name", ""),
            row.get("email", ""),
            row.get("branch_code", ""),
            row.get("branch_name", ""),
            row.get("rate_tier", "standard"),
            row.get("days_worked", 0),
            row.get("regular_hours", 0),
            row.get("overtime_hours", 0),
            row.get("total_hours", 0),
            start_date,
            end_date
        ])
    
    # Add summary row
    total_regular = sum(r.get("regular_hours", 0) for r in data)
    total_ot = sum(r.get("overtime_hours", 0) for r in data)
    total_all = sum(r.get("total_hours", 0) for r in data)
    
    writer.writerow([])
    writer.writerow([
        "TOTAL", "", "", "", "", "", "",
        len(data),
        round(total_regular, 2),
        round(total_ot, 2),
        round(total_all, 2),
        "", ""
    ])
    
    # Audit log
    await audit_log(
        db=db,
        actor_id=current_user.user_id,
        actor_email=current_user.email,
        actor_role=current_user.role,
        action="report.export",
        target_type="payroll",
        target_id="csv",
        metadata={
            "start_date": start_date,
            "end_date": end_date,
            "branch_id": branch_id,
            "employee_count": len(data),
            "total_hours": total_all
        }
    )
    
    output.seek(0)
    filename = f"payroll_{start_date}_to_{end_date}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/attendance-report/csv")
async def export_attendance_report(
    branch_id: Optional[str] = None,
    date: Optional[str] = None,
    current_user: TokenData = Depends(get_current_user)
):
    """Export daily attendance report CSV."""
    db = get_database()
    
    if not has_permission(current_user.role, "reports.export"):
        raise ForbiddenException("Cannot export data")
    
    target_date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    scope = get_role_data_scope(current_user.role)
    
    # Build filter for workers
    user_filter = {"role": "WORKER", "status": "active"}
    if scope == DataScope.BRANCH:
        user_filter["branch_id"] = current_user.branch_id
    elif branch_id and scope == DataScope.ALL:
        user_filter["branch_id"] = branch_id
    
    # Get all workers
    workers = await db.users.find(
        user_filter,
        {"_id": 0, "id": 1, "employee_id": 1, "first_name": 1, "last_name": 1, "email": 1, "branch_id": 1}
    ).to_list(1000)
    
    # Get time entries for the date
    worker_ids = [w["id"] for w in workers]
    entries = await db.time_entries.find(
        {"date": target_date, "user_id": {"$in": worker_ids}},
        {"_id": 0}
    ).to_list(1000)
    
    entries_by_user = {e["user_id"]: e for e in entries}
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        "Employee ID",
        "Name",
        "Email",
        "Status",
        "Clock In",
        "Clock Out",
        "Total Hours",
        "Late",
        "Overtime"
    ])
    
    for worker in workers:
        entry = entries_by_user.get(worker["id"])
        
        if entry:
            status = "Present"
            clock_in = format_time(entry.get("clock_in", {}).get("timestamp"))
            clock_out = format_time(entry.get("clock_out", {}).get("timestamp"))
            total_hours = entry.get("total_hours", 0) or 0
            is_late = entry.get("flags", {}).get("late_clock_in", False)
            ot_hours = entry.get("overtime_hours", 0) or 0
        else:
            status = "Absent"
            clock_in = ""
            clock_out = ""
            total_hours = 0
            is_late = False
            ot_hours = 0
        
        writer.writerow([
            worker.get("employee_id", ""),
            f"{worker.get('first_name', '')} {worker.get('last_name', '')}",
            worker.get("email", ""),
            status,
            clock_in,
            clock_out,
            round(total_hours, 2),
            "Yes" if is_late else "No",
            round(ot_hours, 2)
        ])
    
    output.seek(0)
    filename = f"attendance_{target_date}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@router.get("/payroll/excel")
async def export_payroll_excel(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: TokenData = Depends(get_current_user)
):
    """Export payroll-ready Excel file with formatting."""
    if not HAS_OPENPYXL:
        raise BadRequestException("Excel export not available. openpyxl not installed.")

    db = get_database()

    if not has_permission(current_user.role, "reports.export"):
        raise ForbiddenException("Cannot export data")

    # Default date range
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start = datetime.now(timezone.utc) - timedelta(days=14)
        start_date = start.strftime("%Y-%m-%d")

    scope = get_role_data_scope(current_user.role)

    # Build match filter
    match_filter = {
        "date": {"$gte": start_date, "$lte": end_date},
        "status": {"$in": ["completed", "approved"]}
    }

    if scope == DataScope.BRANCH:
        match_filter["branch_id"] = current_user.branch_id
    elif branch_id and scope == DataScope.ALL:
        match_filter["branch_id"] = branch_id

    # Aggregate by employee
    pipeline = [
        {"$match": match_filter},
        {
            "$group": {
                "_id": "$user_id",
                "total_hours": {"$sum": {"$ifNull": ["$total_hours", 0]}},
                "regular_hours": {"$sum": {"$ifNull": ["$regular_hours", 0]}},
                "overtime_hours": {"$sum": {"$ifNull": ["$overtime_hours", 0]}},
                "days_worked": {"$addToSet": "$date"},
                "entries_count": {"$sum": 1}
            }
        },
        {
            "$lookup": {
                "from": "users",
                "localField": "_id",
                "foreignField": "id",
                "as": "user"
            }
        },
        {"$unwind": "$user"},
        {
            "$lookup": {
                "from": "branches",
                "localField": "user.branch_id",
                "foreignField": "id",
                "as": "branch"
            }
        },
        {"$unwind": {"path": "$branch", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": 0,
                "employee_id": "$user.employee_id",
                "first_name": "$user.first_name",
                "last_name": "$user.last_name",
                "email": "$user.email",
                "branch_code": "$branch.code",
                "branch_name": "$branch.name",
                "rate_tier": "$user.hourly_rate_tier",
                "total_hours": {"$round": ["$total_hours", 2]},
                "regular_hours": {"$round": ["$regular_hours", 2]},
                "overtime_hours": {"$round": ["$overtime_hours", 2]},
                "days_worked": {"$size": "$days_worked"},
                "entries_count": 1
            }
        },
        {"$sort": {"branch_code": 1, "employee_id": 1}}
    ]

    data = await db.time_entries.aggregate(pipeline).to_list(1000)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Report"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    subtitle_font = Font(name="Calibri", size=10, color="666666")
    total_font = Font(name="Calibri", bold=True, size=11)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Title section
    ws.merge_cells("A1:M1")
    ws["A1"] = "PAYROLL REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Pay Period: {start_date} to {end_date} | Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = subtitle_font

    # Headers (row 4)
    headers = [
        "Employee ID", "First Name", "Last Name", "Email",
        "Branch Code", "Branch Name", "Rate Tier",
        "Days Worked", "Regular Hours", "Overtime Hours",
        "Total Hours", "Pay Period Start", "Pay Period End"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(data, 5):
        values = [
            row_data.get("employee_id", ""),
            row_data.get("first_name", ""),
            row_data.get("last_name", ""),
            row_data.get("email", ""),
            row_data.get("branch_code", ""),
            row_data.get("branch_name", ""),
            row_data.get("rate_tier", "standard"),
            row_data.get("days_worked", 0),
            row_data.get("regular_hours", 0),
            row_data.get("overtime_hours", 0),
            row_data.get("total_hours", 0),
            start_date,
            end_date
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in (8, 9, 10, 11):  # Numeric columns
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "0.00"

    # Summary row
    summary_row = len(data) + 5
    total_regular = sum(r.get("regular_hours", 0) for r in data)
    total_ot = sum(r.get("overtime_hours", 0) for r in data)
    total_all = sum(r.get("total_hours", 0) for r in data)

    ws.cell(row=summary_row, column=1, value="TOTAL").font = total_font
    ws.cell(row=summary_row, column=8, value=len(data)).font = total_font
    ws.cell(row=summary_row, column=9, value=round(total_regular, 2)).font = total_font
    ws.cell(row=summary_row, column=10, value=round(total_ot, 2)).font = total_font
    ws.cell(row=summary_row, column=11, value=round(total_all, 2)).font = total_font

    for col in range(1, 14):
        cell = ws.cell(row=summary_row, column=col)
        cell.fill = total_fill
        cell.border = thin_border

    # Column widths
    col_widths = [14, 14, 14, 28, 12, 18, 12, 12, 14, 14, 12, 14, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A5"

    # Audit log
    await audit_log(
        db=db,
        actor_id=current_user.user_id,
        actor_email=current_user.email,
        actor_role=current_user.role,
        action="report.export",
        target_type="payroll",
        target_id="excel",
        metadata={
            "start_date": start_date,
            "end_date": end_date,
            "branch_id": branch_id,
            "employee_count": len(data),
            "total_hours": total_all
        }
    )

    # Write to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"payroll_{start_date}_to_{end_date}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/timesheet/excel")
async def export_timesheet_excel(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = Query(None),
    current_user: TokenData = Depends(get_current_user)
):
    """Export timesheet data as Excel."""
    if not HAS_OPENPYXL:
        raise BadRequestException("Excel export not available.")

    db = get_database()

    if not has_permission(current_user.role, "reports.export"):
        raise ForbiddenException("Cannot export data")

    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start = datetime.now(timezone.utc) - timedelta(days=14)
        start_date = start.strftime("%Y-%m-%d")

    scope = get_role_data_scope(current_user.role)
    status_filter = status.split(",") if status else ["completed", "approved"]

    data = await get_export_data(
        db, branch_id, start_date, end_date, scope,
        current_user.branch_id, current_user.team_id, status_filter
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title
    ws.merge_cells("A1:O1")
    ws["A1"] = f"TIMESHEET REPORT | {start_date} to {end_date}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="2E75B6")

    # Headers
    headers = [
        "Employee ID", "First Name", "Last Name", "Email",
        "Branch Code", "Branch Name", "Date", "Clock In",
        "Clock Out", "Total Hours", "Regular Hours",
        "Overtime Hours", "Break (min)", "Status", "Manual Entry"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, row_data in enumerate(data, 4):
        values = [
            row_data.get("employee_id", ""),
            row_data.get("first_name", ""),
            row_data.get("last_name", ""),
            row_data.get("email", ""),
            row_data.get("branch_code", ""),
            row_data.get("branch_name", ""),
            row_data.get("date", ""),
            format_time(row_data.get("clock_in_time")),
            format_time(row_data.get("clock_out_time")),
            row_data.get("total_hours", 0),
            row_data.get("regular_hours", 0),
            row_data.get("overtime_hours", 0),
            row_data.get("break_minutes", 0),
            row_data.get("status", ""),
            "Yes" if row_data.get("is_manual_entry") else "No"
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    # Auto column widths
    for col in range(1, 16):
        ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = 15

    ws.freeze_panes = "A4"

    await audit_log(
        db=db,
        actor_id=current_user.user_id,
        actor_email=current_user.email,
        actor_role=current_user.role,
        action="report.export",
        target_type="timesheet",
        target_id="excel",
        metadata={"start_date": start_date, "end_date": end_date, "record_count": len(data)}
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"timesheet_{start_date}_to_{end_date}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
